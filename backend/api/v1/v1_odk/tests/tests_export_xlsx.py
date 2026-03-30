import os
from unittest.mock import patch

from django.test import TestCase
from django.test.utils import override_settings
from openpyxl import load_workbook

from api.v1.v1_odk.export import (
    _build_farmer_headers,
    _extract_avg_altitude,
    _get_kobo_base_url,
    _resolve_attachment_url,
    _wkt_to_kml,
    generate_xlsx,
)
from api.v1.v1_odk.models import (
    Farmer,
    FarmerFieldMapping,
    FieldMapping,
    FieldSettings,
    FormMetadata,
    FormQuestion,
    Plot,
    Submission,
)
from api.v1.v1_odk.tests.mixins import (
    OdkTestHelperMixin,
)
from api.v1.v1_users.models import SystemUser
from utils.storage import get_path

VALID_WKT = (
    "POLYGON(("
    "38.47 7.05,"
    "38.48 7.05,"
    "38.48 7.06,"
    "38.47 7.06,"
    "38.47 7.05))"
)

ODK_GEOSHAPE = (
    "7.05 38.47 2980 5;"
    "7.05 38.48 2985 5;"
    "7.06 38.48 2990 5;"
    "7.06 38.47 2985 5;"
    "7.05 38.47 2980 5"
)

EXPORT_URL = "/api/v1/odk/plots/export/"


def _setup_form():
    """Create form with questions, field settings,
    field mappings, and farmer field mapping."""
    form = FormMetadata.objects.create(
        asset_uid="xlsx-form",
        name="XLSX Test Form",
        polygon_field="geoshape",
        region_field="region",
        sub_region_field="woreda",
    )
    q_fname = FormQuestion.objects.create(
        form=form,
        name="First_Name",
        label="First name",
        type="text",
    )
    q_father = FormQuestion.objects.create(
        form=form,
        name="Father_s_Name",
        label="Father name",
        type="text",
    )
    q_gfather = FormQuestion.objects.create(
        form=form,
        name="Grandfather_s_Name",
        label="Grandfather name",
        type="text",
    )
    q_age = FormQuestion.objects.create(
        form=form,
        name="age_of_farmer",
        label="Age",
        type="integer",
    )
    q_td1 = FormQuestion.objects.create(
        form=form,
        name="title_deed_front",
        label="Title Deed Front",
        type="image",
    )
    q_td2 = FormQuestion.objects.create(
        form=form,
        name="title_deed_back",
        label="Title Deed Back",
        type="image",
    )
    FormQuestion.objects.create(
        form=form,
        name="geoshape",
        label="Geoshape",
        type="geoshape",
    )
    FormQuestion.objects.create(
        form=form,
        name="region",
        label="Region",
        type="select_one",
    )
    FormQuestion.objects.create(
        form=form,
        name="woreda",
        label="Woreda",
        type="select_one",
    )

    # Field settings + mappings for title deeds
    fs_td1 = FieldSettings.objects.create(
        name="title_deed_1"
    )
    fs_td2 = FieldSettings.objects.create(
        name="title_deed_2"
    )
    FieldMapping.objects.create(
        field=fs_td1,
        form=form,
        form_question=q_td1,
    )
    FieldMapping.objects.create(
        field=fs_td2,
        form=form,
        form_question=q_td2,
    )

    # Farmer field mapping
    FarmerFieldMapping.objects.create(
        form=form,
        unique_fields=(
            "First_Name,"
            "Father_s_Name,"
            "Grandfather_s_Name"
        ),
        values_fields=(
            "First_Name,"
            "Father_s_Name,"
            "Grandfather_s_Name,"
            "age_of_farmer"
        ),
    )

    return {
        "form": form,
        "q_fname": q_fname,
        "q_father": q_father,
        "q_gfather": q_gfather,
        "q_age": q_age,
        "q_td1": q_td1,
        "q_td2": q_td2,
    }


def _create_plot(
    form,
    kobo_id,
    raw_data,
    farmer=None,
    wkt=VALID_WKT,
):
    """Create a submission + plot pair."""
    sub = Submission.objects.create(
        uuid=f"uuid-{kobo_id}",
        form=form,
        kobo_id=kobo_id,
        submission_time=1700000000000,
        raw_data=raw_data,
    )
    return Plot.objects.create(
        submission=sub,
        form=form,
        plot_name=raw_data.get("First_Name", ""),
        polygon_wkt=wkt,
        min_lat=7.05,
        max_lat=7.06,
        min_lon=38.47,
        max_lon=38.48,
        region="Amhara",
        sub_region="Bahir Dar",
        created_at=1700000000000,
        farmer=farmer,
    )


@override_settings(
    USE_TZ=False, TEST_ENV=True
)
class ExtractAvgAltitudeTest(TestCase):
    def test_valid_geoshape(self):
        form = FormMetadata.objects.create(
            asset_uid="alt-form",
            polygon_field="geoshape",
        )
        raw = {"geoshape": ODK_GEOSHAPE}
        alt = _extract_avg_altitude(raw, form)
        self.assertEqual(alt, 2984.0)

    def test_no_polygon_field(self):
        form = FormMetadata.objects.create(
            asset_uid="alt-form2",
            polygon_field="",
        )
        raw = {"geoshape": ODK_GEOSHAPE}
        alt = _extract_avg_altitude(raw, form)
        self.assertEqual(alt, "")

    def test_missing_geoshape_data(self):
        form = FormMetadata.objects.create(
            asset_uid="alt-form3",
            polygon_field="geoshape",
        )
        raw = {}
        alt = _extract_avg_altitude(raw, form)
        self.assertEqual(alt, "")

    def test_invalid_geoshape(self):
        form = FormMetadata.objects.create(
            asset_uid="alt-form4",
            polygon_field="geoshape",
        )
        raw = {"geoshape": "not valid data"}
        alt = _extract_avg_altitude(raw, form)
        self.assertEqual(alt, "")


@override_settings(
    USE_TZ=False, TEST_ENV=True
)
class ResolveAttachmentUrlTest(TestCase):
    def test_resolves_url(self):
        raw = {
            "title_deed_front": "photo.jpg",
            "_attachments": [
                {
                    "media_file_basename": (
                        "photo.jpg"
                    ),
                    "filename": (
                        "user/attachments/"
                        "abc/photo.jpg"
                    ),
                }
            ],
        }
        url = _resolve_attachment_url(
            raw,
            "title_deed_front",
            "https://kc.kobotoolbox.org",
        )
        self.assertIn(
            "/media/original?media_file=", url
        )
        self.assertIn("photo.jpg", url)

    def test_no_matching_attachment(self):
        raw = {
            "title_deed_front": "missing.jpg",
            "_attachments": [
                {
                    "media_file_basename": (
                        "other.jpg"
                    ),
                    "filename": (
                        "user/attachments/"
                        "abc/other.jpg"
                    ),
                }
            ],
        }
        url = _resolve_attachment_url(
            raw,
            "title_deed_front",
            "https://kc.kobotoolbox.org",
        )
        self.assertEqual(url, "")

    def test_no_field_name(self):
        url = _resolve_attachment_url(
            {}, None, "https://kc.kobotoolbox.org"
        )
        self.assertEqual(url, "")

    def test_empty_field_value(self):
        raw = {"title_deed_front": ""}
        url = _resolve_attachment_url(
            raw,
            "title_deed_front",
            "https://kc.kobotoolbox.org",
        )
        self.assertEqual(url, "")

    def test_no_kobo_url(self):
        raw = {
            "title_deed_front": "photo.jpg",
            "_attachments": [
                {
                    "media_file_basename": (
                        "photo.jpg"
                    ),
                    "filename": (
                        "user/att/photo.jpg"
                    ),
                }
            ],
        }
        url = _resolve_attachment_url(
            raw, "title_deed_front", ""
        )
        self.assertEqual(url, "")


@override_settings(
    USE_TZ=False, TEST_ENV=True
)
class GetKoboBaseUrlTest(TestCase):
    def test_returns_url(self):
        SystemUser.objects.create_superuser(
            email="test@test.org",
            password="pass",
            name="test",
            kobo_url="https://kc.kobotoolbox.org",
            kobo_username="user",
        )
        url = _get_kobo_base_url()
        self.assertEqual(
            url, "https://kc.kobotoolbox.org"
        )

    def test_returns_empty_when_none(self):
        url = _get_kobo_base_url()
        self.assertEqual(url, "")


@override_settings(
    USE_TZ=False, TEST_ENV=True
)
class BuildFarmerHeadersTest(TestCase):
    def test_no_mapping(self):
        form = FormMetadata.objects.create(
            asset_uid="hdr-form",
        )
        headers, fields = (
            _build_farmer_headers(form)
        )
        self.assertEqual(
            headers, ["FarmerID(primary key)"]
        )
        self.assertEqual(fields, [])

    def test_unique_before_values(self):
        ctx = _setup_form()
        form = ctx["form"]
        headers, fields = (
            _build_farmer_headers(form)
        )
        # FarmerID + unique(3) + age(1 extra)
        self.assertEqual(len(headers), 5)
        self.assertEqual(
            headers[0], "FarmerID(primary key)"
        )
        # unique_fields come first
        self.assertEqual(
            headers[1], "First name"
        )
        self.assertEqual(
            headers[2], "Father name"
        )
        self.assertEqual(
            headers[3], "Grandfather name"
        )
        # Then values_fields not in unique
        self.assertEqual(headers[4], "Age")

    def test_deduplicates_fields(self):
        ctx = _setup_form()
        form = ctx["form"]
        _, fields = _build_farmer_headers(form)
        # unique has 3, values has 4 but 3 overlap
        self.assertEqual(len(fields), 4)
        self.assertEqual(
            fields,
            [
                "First_Name",
                "Father_s_Name",
                "Grandfather_s_Name",
                "age_of_farmer",
            ],
        )


@override_settings(
    USE_TZ=False, TEST_ENV=True
)
class GenerateXlsxTest(TestCase):
    def setUp(self):
        self.ctx = _setup_form()
        self.form = self.ctx["form"]

        # Create farmers and plots
        self.farmer1 = Farmer.objects.create(
            uid="00001",
            lookup_key="Dara - Hora - Daye",
            values={
                "First_Name": "Dara",
                "Father_s_Name": "Hora",
                "Grandfather_s_Name": "Daye",
                "age_of_farmer": "38",
            },
        )
        self.farmer2 = Farmer.objects.create(
            uid="00002",
            lookup_key="Dawit - Samuel - Hora",
            values={
                "First_Name": "Dawit",
                "Father_s_Name": "Samuel",
                "Grandfather_s_Name": "Hora",
                "age_of_farmer": "41",
            },
        )
        self.plot1 = _create_plot(
            self.form,
            "101",
            {
                "First_Name": "Dara",
                "Father_s_Name": "Hora",
                "Grandfather_s_Name": "Daye",
                "age_of_farmer": "38",
                "geoshape": ODK_GEOSHAPE,
                "title_deed_front": "td1.jpg",
                "_attachments": [
                    {
                        "media_file_basename": (
                            "td1.jpg"
                        ),
                        "filename": (
                            "u/att/abc/td1.jpg"
                        ),
                    }
                ],
            },
            farmer=self.farmer1,
        )
        self.plot2 = _create_plot(
            self.form,
            "102",
            {
                "First_Name": "Dawit",
                "Father_s_Name": "Samuel",
                "Grandfather_s_Name": "Hora",
                "age_of_farmer": "41",
                "geoshape": ODK_GEOSHAPE,
            },
            farmer=self.farmer2,
        )

    def _generate_and_load(self):
        qs = Plot.objects.filter(form=self.form)
        rel_path, count = generate_xlsx(
            qs, self.form, "test_xlsx"
        )
        full = get_path(rel_path)
        wb = load_workbook(full)
        os.remove(full)
        return wb, count

    def test_produces_valid_xlsx(self):
        wb, count = self._generate_and_load()
        self.assertEqual(count, 2)
        self.assertIn("Farmer table", wb.sheetnames)
        self.assertIn("Plot Table", wb.sheetnames)

    def test_farmer_sheet_headers(self):
        wb, _ = self._generate_and_load()
        ws = wb["Farmer table"]
        headers = [
            c.value for c in ws[1]
        ]
        self.assertEqual(
            headers[0], "FarmerID(primary key)"
        )
        self.assertEqual(
            headers[1], "First name"
        )
        self.assertEqual(
            headers[2], "Father name"
        )

    def test_farmer_sheet_one_row_per_farmer(self):
        wb, _ = self._generate_and_load()
        ws = wb["Farmer table"]
        # header + 2 farmers
        self.assertEqual(ws.max_row, 3)

    def test_farmer_id_prefix(self):
        wb, _ = self._generate_and_load()
        ws = wb["Farmer table"]
        ids = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
        ]
        self.assertIn("AB00001", ids)
        self.assertIn("AB00002", ids)

    def test_farmer_values_populated(self):
        wb, _ = self._generate_and_load()
        ws = wb["Farmer table"]
        # Find AB00001 row
        for row in ws.iter_rows(
            min_row=2, values_only=True
        ):
            if row[0] == "AB00001":
                self.assertEqual(row[1], "Dara")
                self.assertEqual(row[2], "Hora")
                self.assertEqual(row[3], "Daye")
                self.assertEqual(row[4], "38")
                break

    def test_plot_sheet_headers(self):
        wb, _ = self._generate_and_load()
        ws = wb["Plot Table"]
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers[0], "Submission ID")
        self.assertEqual(headers[1], "Farmer ID")
        self.assertEqual(
            headers[2], "Title Deed First Page"
        )

    def test_plot_sheet_one_row_per_plot(self):
        wb, _ = self._generate_and_load()
        ws = wb["Plot Table"]
        # header + 2 plots
        self.assertEqual(ws.max_row, 3)

    def test_submission_id_prefix(self):
        wb, _ = self._generate_and_load()
        ws = wb["Plot Table"]
        ids = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
        ]
        self.assertIn("#101", ids)
        self.assertIn("#102", ids)

    def test_plot_references_farmer(self):
        wb, _ = self._generate_and_load()
        ws = wb["Plot Table"]
        farmer_ids = [
            ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
        ]
        self.assertIn("AB00001", farmer_ids)
        self.assertIn("AB00002", farmer_ids)

    def test_altitude_from_geoshape(self):
        wb, _ = self._generate_and_load()
        ws = wb["Plot Table"]
        # Column 7 = Altitude
        alt = ws.cell(row=2, column=7).value
        self.assertIsNotNone(alt)
        self.assertAlmostEqual(alt, 2984.0)

    def test_centroid_coordinates(self):
        wb, _ = self._generate_and_load()
        ws = wb["Plot Table"]
        # Column 5 = Latitude, 6 = Longitude
        lat = ws.cell(row=2, column=5).value
        lon = ws.cell(row=2, column=6).value
        self.assertIsNotNone(lat)
        self.assertIsNotNone(lon)
        self.assertAlmostEqual(
            lat, 7.055, places=2
        )
        self.assertAlmostEqual(
            lon, 38.475, places=2
        )

    def test_polygon_kml_url(self):
        wb, _ = self._generate_and_load()
        ws = wb["Plot Table"]
        url = ws.cell(row=2, column=8).value
        self.assertIn(
            "/api/v1/odk/plots/", url
        )
        self.assertIn("/kml/?key=", url)

    def test_no_geometry_still_included(self):
        """Plots without polygon should still
        appear in XLSX export."""
        _create_plot(
            self.form,
            "103",
            {"First_Name": "NoGeo"},
            wkt=None,
        )
        qs = Plot.objects.filter(form=self.form)
        _, count = generate_xlsx(
            qs, self.form, "test_nogeo"
        )
        self.assertEqual(count, 3)


@override_settings(
    USE_TZ=False, TEST_ENV=True
)
class ExportXlsxEndpointTest(
    TestCase, OdkTestHelperMixin
):
    def setUp(self):
        self.user = self.create_kobo_user()
        self.auth = self.get_auth_header()
        self.ctx = _setup_form()

    @patch(
        "api.v1.v1_odk.views.async_task",
        return_value="fake-task-id",
    )
    def test_export_xlsx_creates_job(
        self, _mock_task
    ):
        resp = self.client.post(
            EXPORT_URL,
            {
                "form_id": "xlsx-form",
                "format": "xlsx",
            },
            content_type="application/json",
            **self.auth,
        )
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertEqual(
            data["type"], "export_xlsx"
        )
        self.assertEqual(
            data["status"], "pending"
        )

    @patch(
        "api.v1.v1_odk.views.async_task",
        return_value="fake-task-id",
    )
    def test_xlsx_format_accepted(
        self, _mock_task
    ):
        resp = self.client.post(
            EXPORT_URL,
            {
                "form_id": "xlsx-form",
                "format": "xlsx",
            },
            content_type="application/json",
            **self.auth,
        )
        self.assertEqual(resp.status_code, 201)


@override_settings(
    USE_TZ=False, TEST_ENV=True
)
class WktToKmlTest(TestCase):
    def test_valid_polygon(self):
        kml = _wkt_to_kml(VALID_WKT, name="Test")
        self.assertIn("<kml", kml)
        self.assertIn("<name>Test</name>", kml)
        self.assertIn("<coordinates>", kml)
        self.assertIn("38.47,7.05,0", kml)

    def test_empty_string(self):
        self.assertEqual(_wkt_to_kml(""), "")

    def test_none(self):
        self.assertEqual(_wkt_to_kml(None), "")

    def test_invalid_wkt(self):
        self.assertEqual(
            _wkt_to_kml("INVALID"), ""
        )
